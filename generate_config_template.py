# -*- coding: utf-8 -*-
"""生成部门结算任务配置模板 Excel 文件"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE_DIR, '部门结算任务配置模板.xlsx')

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '结算任务配置'

    # 样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='微软雅黑', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 标题行
    headers = ['结算类型', '部门名称', '需提供资料', '截止日期(每月几日前)', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 示例数据（来自管理办法文档）
    data = [
        # 对上结算
        ('对上结算', '生产调度部', '当月完成工程量报表、施工日志汇总、借用设备确认单（与物资部确认是否有使用联营体零星机械）', 28, '结算基础依据'),
        ('对上结算', '工程质量技术部', '质量验收记录、监理验工资料、对应图纸、专项计量资料、图纸工程量对比表、计量计算底稿等', 29, '不合格不予结算'),
        ('对上结算', '安全环保部', '总包安全检查整改回执、HSE罚款', 29, '涉及安全费扣款/计取'),
        ('对上结算', '设备物资部', '分包商材料领用表（甲供材消耗统计表）、安全领用材料、设备使用扣款单、材料调拨单', 30, '扣回甲供材费用'),
        # 2026-08-13：HR 并入综合办公室，HR 的农民工工资任务由综合办公室承接
        ('对上结算', '综合办公室', '农民工工资发放表、考勤表、分包人员宿舍租赁、水电、食堂伙食费扣款单', 30, '证明无欠薪 + 总承包管理费分摊'),
        ('对上结算', '财务资金部', '已收款台账、发票开具情况、税金预缴凭证', 30, '核对累计结算金额'),
        # 对下结算
        ('对下结算', '生产调度部', '分包当月完成工程量核定表，借用设备确认单', 28, '核量控制（以生产部为准），防止虚报'),
        ('对下结算', '工程质量技术部', '分包工程验收记录、质量扣款单', 28, '质量扣款依据，不合格不结算'),
        ('对下结算', '安全环保部', '安全违规处罚单、文明施工考核扣分', 28, '安全扣款依据'),
        ('对下结算', '设备物资部', '甲供材领用台账、超耗扣款表、分包代采物资计量单', 28, '材料扣款'),
        # 2026-08-13：HR 并入综合办公室，工资代发/考勤任务由综合办公室承接
        ('对下结算', '综合办公室', '工资代发清单（即分包用菲籍工人）、考勤抽查记录、自用人工费、水电、住宿、食堂等后勤扣款', 28, '工资发放核查 + 现场经费扣款'),
        ('对下结算', '财务资金部', '已付款台账、预付款扣回情况', 28, '付款控制，防止超付'),
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # 列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25

    # 行高
    ws.row_dimensions[1].height = 30
    for r in range(2, len(data) + 2):
        ws.row_dimensions[r].height = 35

    # 说明 sheet
    ws2 = wb.create_sheet('填写说明')
    instructions = [
        ('部门结算任务配置文件 - 填写说明', ''),
        ('', ''),
        ('1. 结算类型', '填写"对上结算"或"对下结算"'),
        ('2. 部门名称', '填写部门全称，如：生产调度部、工程质量技术部等'),
        ('3. 需提供资料', '详细描述该部门需要提供/审核的资料内容'),
        ('4. 截止日期', '填写数字，表示每月几日前提交（如28表示每月28日前）'),
        ('5. 备注', '选填，补充说明'),
        ('', ''),
        ('注意：', ''),
        ('- 系统会自动识别标题行，请勿修改第一行标题', ''),
        ('- 如需新增部门，直接在表中添加新行即可', ''),
        ('- 上传后将更新系统中的部门及任务配置', ''),
        ('- 已有配置会被覆盖更新，不会删除', ''),
    ]
    for row_idx, (label, desc) in enumerate(instructions, 1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(name='微软雅黑', size=11, bold=(row_idx <= 1 or label.endswith('：')))
        ws2.cell(row=row_idx, column=2, value=desc).font = Font(name='微软雅黑', size=10)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 60

    wb.save(OUTPUT)
    print(f'配置模板已生成: {OUTPUT}')


if __name__ == '__main__':
    main()
